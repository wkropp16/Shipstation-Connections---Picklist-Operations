import pypyodbc, ShipStation, datetime, json, time, MySQLdb, _mysql, shutil, os, GoogleConnection, SendEmail, pyodbc
from operator import itemgetter
from openpyxl import *
from openpyxl.utils import *
from openpyxl.styles import *
from openpyxl.worksheet import header_footer
from datetime import datetime as dt
from datetime import timedelta
from MySQLdb import converters

def main():
	conv=converters.conversions.copy()
	conv[246]=float	# convert decimals to floats
	conn = MySQLdb.connect(host="localhost", user="REMOVED", passwd="REMOVED", db="shipstation", charset='utf8', conv=conv)
	cursor = conn.cursor()
	ordDate = (dt.now()-timedelta(days=8)).strftime("%Y-%m-%d")
	createPick(conn, cursor, ordDate)
	print('')
	print("INFO - Closing Database Connection . . .")
	conn.commit()
	cursor.close()
	conn.close()
	print("INFO - Database Connection Closed . . .")
	print('')
	SendEmail.sendEmailWithAttachment('Picklist', 'Picklist Created Successfully', 'The Picklist file has been created', 'Picklist')
	return;

def getItemExceptions(): # Gets info from the ItemExceptions Sheet to match against order skus to determine how to split them or make them multi-product skus
	exceptionlist = {}
	xlfile = ('../References/itemexceptions.xlsx') 	# Specify File Name
	wb = load_workbook(filename = xlfile) 			# Define the 'workbook'
	IEMASTER = wb['Master'] 						# Define Sheet
	for row in IEMASTER.iter_rows(min_row=2, max_col=IEMASTER.max_column, max_row=IEMASTER.max_row): # Make a loop for rows
		if row != "":
			alternates = {}
			col_a, col_b, col_c, col_d, col_e, col_f = row	#Splits the 'row' into values
			a_value = col_a.value							#Sold Sku
			b_value = str(col_b.value).split(',')			#Css Sku(s)
			c_value = str(col_c.value).split(',')			#Qty
			if ',' in str(col_e.value):
				e_value = str(col_e.value).split(',')		#Alternate Warehouse
				for i in range(0,len(b_value)):
					alternates[b_value[i]] = (c_value[i],e_value[i])
			else: 
				e_value = str(col_e.value)					#Alt warehouse if non split
				for i in range(0,len(b_value)):
					alternates[b_value[i]] = (c_value[i],e_value)

			exceptionlist[a_value] = alternates
	wb.close()
	return exceptionlist;

def getGrouponExceptions():
	GrouponExceptions = {}
	fileName = GoogleConnection.getSheetInfo('REMOVED', 'Contracted!A1:AG')
	wb = load_workbook(filename = fileName) 	#Define the 'workbook'
	gsMaster = wb['MASTER'] 	#Define Sheet

	for row in gsMaster.iter_rows(min_row=3, max_col=gsMaster.max_column, max_row=gsMaster.max_row):
		if row != "":
			itemdetails = []
			col_a, col_b, col_c, col_d, col_e, col_f, col_g, col_h, col_i, col_j, col_k, col_l, col_m, col_n, col_o, col_p, col_q, col_r, col_s, col_t, col_u, col_v, col_w, col_x, col_y, col_z, col_aa, col_ab, col_ac, col_ad, col_ae, col_af, col_ag = row	#Splits the 'row' into values
			# 0		1		2		3		4	5		6		7		8	9		10		11	12		13		14		15		16	17		18		19		20	21		22		23		24	
			# col_a = Build No # col_j = Groupon UPC # col_ae = Deal Notes # col_n = Ship WHSE # col_l = Root Sku(s) # col_m = Qty per sale # col_k = Shipstation Sku # col_c = Product Name/Item Desc
			gBuildno = col_a.value		# Build No
			gUPC = col_j.value			# Groupon UPC
			gNotes = col_ae.value		# Deal Notes
			gShipWhse = col_n.value		# Ship WHSE
			if col_l.value == 'None' or col_j.value == 'None' or col_h.value == 'None':
				break;
			if col_l.value is not None and ',' in col_l.value:
				gRootSku = str(col_l.value).split(',')					# Root Sku(s)
				gQtyPerSale = str(col_m.value).split(',')				# Qty per sale
				# print('QTY VALUE = ' + str(gQtyPerSale))
				gPickComments = ('Part of kit: ' + str(col_k.value)) 	# Shipstation Sku (makes pick comments)
				for i in range(0,len(gRootSku)): 	# Root Sku(s)
					# print(gRootSku[i],gQtyPerSale[i],gShipWhse,gPickComments,gNotes,gBuildno)
					if str(gNotes) == 'None': 		# Deal Notes
						gNotes = ''					# Deal Notes
					itemdetails.append((gRootSku[i],gQtyPerSale[i],gShipWhse,gPickComments,gNotes,gBuildno)) 	# 'UPC': {'SKU': (QTY, 'SHIPWHSE','Part of Kit', Notes, 'BUILDNO')}
			else:
				gRootSku = col_l.value				# Root Sku(s)
				gQtyPerSale = col_m.value			# Qty per sale
				gPickComments = str(col_c.value)	# Product Name/Item Desc (makes pick comments)
				itemdetails.append((gRootSku, gQtyPerSale,gShipWhse,gPickComments,gNotes,gBuildno)) 			# 'UPC': {'SKU': (QTY, 'SHIPWHSE','ProdName', Notes, 'BUILDNO')}
			GrouponExceptions[gUPC] = itemdetails # Groupon UPC

	print('INFO - Lookup Table Created')
	wb.close()
	return GrouponExceptions;


def getPicklistStores(): # Gets the stores from ../References/EssentialValues.xlsx to determine which stores should be on the picklist
	storelist = []
	xlfile = ('../References/essentialvalues.xlsx') 							# Specify File Name
	wb = load_workbook(filename = xlfile) 										# Define the 'workbook'
	EVPLS = wb['PICKLIST_STORES'] 												# Define Sheet
	for row in EVPLS.iter_rows(min_row=2, max_col=2, max_row=EVPLS.max_row):	# Make a loop for rows
		col_a, col_b = row
		if col_a.value != None:
			storelist.append(col_a.value)
	wb.close()
	return storelist;

def getStoreBins(): # Gets the stores from ../References/EssentialValues.xlsx to determine which stores should be on the picklist
	storeBins = {}
	xlfile = ('../References/essentialvalues.xlsx') 							# Specify File Name
	wb = load_workbook(filename = xlfile) 										# Define the 'workbook'
	EVPLI = wb['PICKLIST_INFO'] 												# Define Sheet
	for row in EVPLI.iter_rows(min_row=2, max_col=2, max_row=EVPLI.max_row):	# Make a loop for rows
		col_a, col_b = row
		if col_a.value != None:
			storeBins[col_a.value] = col_b.value
	wb.close()
	return storeBins;

def getItemLocationDetailsOLD(sku): # Returns BinLocation Values from the Portal's Database
	conn2 = pyodbc.connect(r"REMOVED")
	cursor2 = conn2.cursor()
	CMD_CheckPortal = ('SELECT IM404_ItemLocationQuantity.ItemCode, IM404_ItemLocationQuantity.WarehouseCode, IM404_ItemLocationQuantity.BinLocation, IM404_ItemLocationQuantity.QuantityOnHand, IM404_ItemLocationQuantity.QuantityAllocated, IM404_BinLocation.LocationType FROM IM404_ItemLocationQuantity INNER JOIN IM404_BinLocation ON IM404_ItemLocationQuantity.BinLocation=IM404_BinLocation.BinLocation WHERE IM404_ItemLocationQuantity.ItemCode=?')

	cursor2.execute(CMD_CheckPortal,(sku,))
	locationTuples2 = []
	binList = []
	for location in cursor2.fetchall(): # SQL Command to get all the locations for the sku passed to this definition
		if float(location[3]) > 0 and location[5] == 'P' and location[2] != 'SHIPPING' and location[2] != 'DIGITALDATA': # Check if the location >0 inventory and is Pickable
			location[3] = float(location[3]) # Location QuantityOnHand
			location[4] = float(location[4]) # Location QuantityAllocated
			for alreadyAddedLoc in locationTuples2:
				binList.append(alreadyAddedLoc[2])
			if location[2] in binList: # skip adding the bin to be picked from
				continue;
			else:
				locationTuples2.append(location)
	locationTuples = sorted(locationTuples2, key=itemgetter(3), reverse=True)
	return locationTuples;

def getItemLocationDetails(sku): # Returns BinLocation Values from the Portal's Database
	conn2 = pyodbc.connect(r"REMOVED")
	cursor2 = conn2.cursor()
	CMD_CheckPortal = ('SELECT IM404_ItemLocationQuantity.ItemCode, IM404_ItemLocationQuantity.WarehouseCode, IM404_ItemLocationQuantity.BinLocation, IM404_ItemLocationQuantity.QuantityOnHand, IM404_ItemLocationQuantity.QuantityAllocated, IM404_BinLocation.LocationType FROM IM404_ItemLocationQuantity INNER JOIN IM404_BinLocation ON IM404_ItemLocationQuantity.BinLocation=IM404_BinLocation.BinLocation WHERE IM404_ItemLocationQuantity.ItemCode=?')

	cursor2.execute(CMD_CheckPortal,(sku,))
	warehouseLocations = []
	cageLocations = []
	warehouseBinList = []
	cageBinList = []
	locations = []
	for location in cursor2.fetchall(): # SQL Command to get all the locations for the sku passed to this definition
		if float(location[3]) > 0 and location[5] == 'P' and location[2] != 'SHIPPING' and location[2] != 'DIGITALDATA' and location[2] != 'DOCK' and location[2] != 'CAGE-PICK' and location[2].startswith('CAGE') == False: # Check if the location >0 inventory and is Pickable
			location[3] = float(location[3]) # Location QuantityOnHand
			location[4] = float(location[4]) # Location QuantityAllocated
			for alreadyAddedLoc in warehouseLocations:
				warehouseBinList.append(alreadyAddedLoc[2])
			if location[2] in warehouseBinList: # skip adding the bin to be picked from
				continue;
			else:
				warehouseLocations.append(location)
		elif float(location[3]) > 0 and location[5] == 'P' and location[2].startswith('CAGE') and location[2] != 'CAGE-PICK':
			location[3] = float(location[3]) # Location QuantityOnHand
			location[4] = float(location[4]) # Location QuantityAllocated
			for alreadyAddedLoc in cageLocations:
				cageBinList.append(alreadyAddedLoc[2])
			if location[2] in cageBinList: # skip adding the bin to be picked from
				continue;
			else:
				cageLocations.append(location)
	if cageLocations:
		locations.insert(0, cageLocations[0])
	else:
		locations.insert(0, ())
	if warehouseLocations:
		locations.insert(1, warehouseLocations[0])
	else:
		locations.insert(1, ())
	return locations;

def createPick(conn, cursor, ordDate):
	print('INFO - - - - - - - - - - - - - - - - - Processing Orders! ! !')
	start = time.time()

	print('INFO - Getting Exceptions . . . ')
	itemexceptions = getItemExceptions()
	GrouponExceptions = getGrouponExceptions()
	print('INFO - Exceptions Recieved\n')

	print('INFO - Getting Picklist Stores . . . ')
	stores = getPicklistStores()
	print('INFO - Stores Recieved\n')

	CMD_findOpenOrders = 		('SELECT orderId, orderNumber, orderStatus, orderDate, ao_StoreId, userId FROM ss_orders WHERE orderStatus=%s order by orderDate desc')
	CMD_getOrderItems = 		('SELECT orderId, orderSku, orderSkuName, orderSkuQuantity, orderSkuUnitPrice, orderSkuTaxAmount, orderSkuShippingAmount, lineItemKey FROM ss_order_items WHERE orderId=%s')
	CMD_addToPickList = 		('INSERT INTO picklist (orderId, orderNumber, orderDate, sku, skuIfMisc, orderSkuUnitPrice, quantity, grouponQty, warehouseAssigned, location1, location2, Status, lineItemKey) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)')
	CMD_CheckPicklist = 		('SELECT orderId, orderNumber, sku, quantity, warehouseAssigned, lineItemKey FROM picklist WHERE orderId=%s AND orderNumber=%s AND sku=%s AND quantity=%s and lineItemKey=%s')
	CMD_deleteFromPicklist = 	('DELETE FROM picklist WHERE uniqueId=%s AND orderId=%s')
	CMD_CheckdeleteFromPicklist = ('SELECT orderId, orderNumber, sku, quantity, warehouseAssigned, lineItemKey, uniqueId FROM picklist WHERE orderId=%s AND orderNumber=%s AND lineItemKey=%s')
	
	print('INFO - Getting \'Cancelled\' Orders . . . ')
	cursor.execute(CMD_findOpenOrders, ('cancelled',))
	cancelledOrders = cursor.fetchall()
	print('INFO - Cancelled Orders Recieved')
	for order in cancelledOrders:
		if order[4] in stores:
			cursor.execute(CMD_getOrderItems, (order[0],))
			cancelledItems = cursor.fetchall()
			for item in cancelledItems: # (('395072681', 'HB-9078', 'Ped Egg Callus Remover', Decimal('1.00'), Decimal('14.99'), None, None),)
				cursor.execute(CMD_CheckdeleteFromPicklist, (order[0], order[1], item[7]))
				deleteThisOrder = cursor.fetchall()
				for orda in deleteThisOrder:
					cursor.execute(CMD_deleteFromPicklist, (orda[6], orda[0]))
					print('	INFO - Order:', orda[0], '|', orda[6], 'removed from picklist (cancelled order)')
	print('INFO - Cancelled Orders Processed! ! !\n')



	print('INFO - Getting \'Awaiting Shipment\' Orders . . . ')
	cursor.execute(CMD_findOpenOrders, ('awaiting_shipment',))
	openOrders = cursor.fetchall()
	print('INFO - Awaiting-Shipment Orders Recieved\n')
	print('INFO - Processing Open-Orders to Open-Order-Items and Correcting Skus')

	addToPicklist2 = []
	ordersListforAssign = []

	for order in openOrders: # ('orderId', 'orderNum', 'orderStatus', orderDate, 'orderStore', 'userId')
		if order[4] in stores and order[5] == None: # Check if valid pickable store & if not assigned to anyone
			print('	INFO -	OrderNum', order[1], '	|	OrderID', order[0])
			CMD_2 = ("SELECT * FROM ss_orders where ao_MergedIds like %s")
			likestmnt=str('%' + str(order[0]) + '%')
			cursor.execute(CMD_2, (likestmnt,))
			cmd2 = cursor.fetchone()
			print('	INFO -	CMD2 - ', cmd2)
			if cmd2 is not None:
				print('	INFO -	CMD2 NOT NONE - ', cmd2)
				print('	INFO -	Skipping	', order[1], ' From adding to Picklist (Merged)')
				CMD_3 = ("UPDATE ss_orders SET orderStatus=%s WHERE orderId=%s")
				cursor.execute(CMD_3, ('shipped', order[0]))
				continue;

			cursor.execute(CMD_getOrderItems, (order[0],))
			openItems = cursor.fetchall()

			for item in openItems: # (('395072681', 'HB-9078', 'Ped Egg Callus Remover', Decimal('1.00'), Decimal('14.99'), None, None),)
				# print('	', 'OrderID', item[0], 'Sku', item[1], 'QTY', item[3], 'LineItemID', item[7])
				originalSku = item[1]
				location1 = 'NA'
				location2 = 'NA'
				newSku = str(originalSku)
				skuDesc = str(item[2])[:50]
				newMiscSku = ''
				newQuantity = item[3]
				newPerUnitPrice = item[4]
				newWarehouse = '7'
				grouponQty = 0

				if originalSku in itemexceptions: #Check for itemexceptions
					for exceptionSku, exceptionSkuvalues in itemexceptions[originalSku].items():
						newSku = exceptionSku
						newMiscSku = ''
						newQuantity = float(float(exceptionSkuvalues[0])*item[3])
						newPerUnitPrice = float(item[4]/newQuantity)
						newWarehouse = exceptionSkuvalues[1]

				elif originalSku in GrouponExceptions:
					for piecesList in GrouponExceptions[str(originalSku)]:
						if piecesList[2] == 'Las Vegas':
							newWarehouse = '100'
						else:
							newWarehouse ='7'
						newSku = piecesList[0]
						newQuantity = float(float(piecesList[1]) * item[3])
						newPerUnitPrice = float(item[4]/newQuantity)
						newMiscSku = ''

				elif originalSku == "" or originalSku.startswith("MS-") or originalSku.startswith("TIGER") or originalSku.startswith("MISC-"): #check for MISC items
					newSku = 'MISC'
					newMiscSku = originalSku
					newQuantity = item[3]
					newWarehouse = '7'

				elif originalSku.startswith("TEST"): #check for TEST items
					newSku = 'TEST'
					newMiscSku = originalSku
					newQuantity = item[3]
					newWarehouse = '7'

				if newSku != 'MISC' or newSku != 'TEST':
					skuLocations = getItemLocationDetails(newSku)
					if skuLocations[0]:
						location1 = str('*' + skuLocations[0][2] + '*')
					else:
						location1 = ''

					if skuLocations[1]:
						location2 = str('*' + skuLocations[1][2] + '*')
					else:
						location2 = ''

				if order[4] == 'SP - Groupon':
					grouponQty = newQuantity
				
				cursor.execute(CMD_CheckPicklist, (order[0], order[1], newSku, newQuantity, item[7]))
				itemInPicklist = cursor.fetchone()
				if not itemInPicklist:
					print('Order Item not found in Picklist, creating new entry . . .')
					ordersListforAssign.append([order[0], newWarehouse, order[1]])
					cursor.execute(CMD_addToPickList, (order[0], order[1], order[3], newSku, newMiscSku, newPerUnitPrice, newQuantity, grouponQty, newWarehouse, location1, location2, order[2], item[7]))
					# orderId, orderNumber, orderDate, sku, skuIfMisc, orderSkuUnitPrice, quantity, grouponQty, warehouseAssigned, location1, location2, Status, lineItemKey
					# comment out execute
					conn.commit() ############################

					if not addToPicklist2:
						if str(newWarehouse) == '7':
							addToPicklist2.append([newSku, skuDesc, newQuantity, location1, location2, newMiscSku, grouponQty])
					else:
						for items in addToPicklist2:
							if items[0] == newSku: 										# If picksku=alreadyinsku
								if items[0] == 'MISC' and items[5] != newMiscSku:		# if sku=misc and miscSku!=miscSku
									if str(newWarehouse) == '7':						# If warehouse=7
										addToPicklist2.append([newSku, skuDesc, newQuantity, location1, location2, newMiscSku, grouponQty])
										break
								elif items[0] == 'MISC' and items[5] == newMiscSku:		# if sku=misc and miscSku=miscSku
									items[2] += newQuantity
									items[6] += grouponQty
									break
								items[2] += newQuantity
								items[6] += grouponQty
								break
						else:
							if str(newWarehouse) == '7':
								addToPicklist2.append([newSku, skuDesc, newQuantity, location1, location2, newMiscSku, grouponQty])
								continue
				elif itemInPicklist:
					print('	INFO - Item is already in Picklist  | ', '	', 'OrderID', item[0], 'Sku', item[1], 'QTY', item[3], 'LineItemID', item[7])

	print('INFO - Open-Orders Processed! ! !\n')

	# addToPicklist = sorted(addToPicklist2, key=itemgetter(6,3), reverse=True)
	addToPicklist2.sort(key=itemgetter(6), reverse=True)

	picklistFile(addToPicklist2)
	assignOrders(ordersListforAssign)
	# comment out assigning

	end = time.time()
	# os.startfile('V:\Orders-2.0\currentPicklist.xlsx', 'print') # Un-comment to auto-print picklist
	print('INFO - Completed in: ' + str(end - start)[:14] + ' seconds.\n')
	conn.commit() ########################
	print("INFO - Rows Commited to Database")
	return;

def picklistFile(addToPicklist2):
	#create Workbook for picklist
	wb = Workbook()
	destinationName = '../currentPicklist.xlsx'
	ws = wb.worksheets[0]
	ws.title = 'MAIN'
	sideBarFont = Font(name='IDAutomationSHC39S Demo', size=24)	# https://www.idautomation.com/barcode-fonts/code-39/fontnames.html
	otherFont = Font(name='IDAutomationSHC39L Demo', size=10)	# https://www.idautomation.com/barcode-fonts/code-39/fontnames.html
	sideBarRanges = ['A5:A18', 'A19:A32', 'A33:A46', 'A47:A60']
	headerRanges = ['B1:C1', 'B2:C2', 'B4:C4', 'D1:E1', 'D2:E2', 'D4:F4', 'H1:J1', 'H2:J2', 'H4:J4', 'K4:M4', 'L2:N2', 'N4:O4']
	boldRanges = ['B4', 'D4', 'G4', 'H4', 'K4', 'N4']
	centerAlignRanges = ['D1', 'D2', 'L1', 'B4', 'D4', 'G4', 'H4', 'K4', 'N4']
	boldText = Font(name='Calibri', size=11, bold=True)
	centerAlign = Alignment(vertical='center', horizontal='center', wrap_text=False)
	sidebarAlign = Alignment(vertical='center', horizontal='center', text_rotation=90)
	leftCenterAlign = Alignment(vertical='center', horizontal='left', wrap_text=True)
	rightAlignHoriz = Alignment(horizontal='right')
	colHeaderBorder = Border(top=Side(style='thick'), right=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))
	TB = Border(top=Side(style='thin'), bottom=Side(style='thin'))
	TLB = Border(top=Side(style='thin'), left=Side(style='thin'), bottom=Side(style='thin'))
	TRB = Border(top=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))
	skuWriteRow = 5 # Starting row to add SKUs to the picklist
	shipBin = '*E-FULFILLMENT*' # Side Scancode for bin moval

	# Print Setup
	ws.set_printer_settings(ws.PAPERSIZE_A4, ws.ORIENTATION_LANDSCAPE)
	ws.page_setup.horizontalCentered = True
	ws.page_setup.verticalCentered = True
	ws.print_title_rows = '1:4'
	ws.page_margins.left = 0.25
	ws.page_margins.right = 0.25
	ws.page_margins.top = 0.25
	ws.page_margins.bottom = 0.25
	ws.oddHeader.right.text = "&[Page] of &N"

	# Merge Cells
	ws.merge_cells(range_string=sideBarRanges[0])
	for cellrange in headerRanges:
		ws.merge_cells(range_string=cellrange)

	# Style Cells
	ws.row_dimensions[4].height = 30
	ws['C1'].border = TB
	ws['C2'].border = TB
	ws['D1'].border = TB
	ws['D2'].border = TB
	ws['H1'].border = TB
	ws['I1'].border = TB
	ws['H2'].border = TB
	ws['I2'].border = TB
	ws['M2'].border = TB
	ws['B1'].border = TLB
	ws['B2'].border = TLB
	ws['G1'].border = TLB
	ws['G2'].border = TLB
	ws['L2'].border = TLB
	ws['E1'].border = TRB
	ws['E2'].border = TRB
	ws['J1'].border = TRB
	ws['J2'].border = TRB
	ws['N2'].border = TRB
	ws['D1'].value = str(dt.today().strftime('%a. %b. %d, %Y'))
	ws['D2'].value = str(dt.today().strftime('%I:%M %p'))
	ws['B1'].value = 'Picklist Date'
	ws['B2'].value = 'Picklist Time'
	ws['G1'].value = 'Pick By'
	ws['G2'].value = 'Pack By'
	ws['L2'].value = 'Ecommerce Picklist'
	ws['L2'].alignment = centerAlign
	ws['B4'].value = 'SKU'
	ws['D4'].value = 'Description'
	ws['G4'].value = 'QTY'
	ws['H4'].value = 'Cage Location'
	ws['K4'].value = 'Warehouse Location'
	ws['N4'].value = 'Sku ScanCode'
	ws['B1'].alignment = rightAlignHoriz
	ws['B2'].alignment = rightAlignHoriz

	for cellval in boldRanges:
		ws[cellval].font = boldText
	for cellval in centerAlignRanges:
		ws[cellval].alignment = centerAlign
	for row in ws.iter_rows(min_row=4, max_row=4, max_col=15):
		for cell in row:
			cell.border = colHeaderBorder
	for i, col in enumerate(ws.iter_cols(min_col=None, max_col=15)): # Set Column Widths
		widths = [6.28, 9.14, 9.14, 9.14, 9.14, 9.14, 7.00, 8.42, 8.42, 8.42, 8.42, 8.42, 8.42, 16.14, 16.14]
		ws.column_dimensions[get_column_letter(i + 1)].width = widths[i]
	for i, row in enumerate(ws.iter_rows(min_row=None, max_row=4)): # Set Row Heights
		ws.row_dimensions[(i + 1)].height = 20
	# for i, row in enumerate(ws.iter_rows(min_row=5, max_row=100)): # Set Row Heights
	# 	ws.row_dimensions[(i + 5)].height = 35

	for item in addToPicklist2:
		lineItemMerges = []
		lineItemMerges.append("B"+str(skuWriteRow)+":C"+str(skuWriteRow)+"")
		lineItemMerges.append("D"+str(skuWriteRow)+":F"+str(skuWriteRow)+"")
		lineItemMerges.append("H"+str(skuWriteRow)+":J"+str(skuWriteRow)+"")
		lineItemMerges.append("K"+str(skuWriteRow)+":M"+str(skuWriteRow)+"")
		lineItemMerges.append("N"+str(skuWriteRow)+":O"+str(skuWriteRow)+"")
		for lineItemMerge in lineItemMerges:
			ws.merge_cells(range_string=lineItemMerge)

		ws.row_dimensions[skuWriteRow].height = 35

		ws['G4'].value = 'QTY'
		ws["B"+str(skuWriteRow)].value = str(item[0]) # Sku
		ws["B"+str(skuWriteRow)].alignment = centerAlign

		ws["D"+str(skuWriteRow)].value = str(item[5] + item[1]) # Description
		ws["D"+str(skuWriteRow)].alignment = leftCenterAlign

		ws["G"+str(skuWriteRow)].value = str(item[2]) # Quantity
		ws["G"+str(skuWriteRow)].alignment = centerAlign

		ws["H"+str(skuWriteRow)].value = str(item[3]) # Location 1
		ws["H"+str(skuWriteRow)].font = otherFont
		ws["H"+str(skuWriteRow)].alignment = centerAlign

		ws["K"+str(skuWriteRow)].value = str(item[4]) # Location 2
		ws["K"+str(skuWriteRow)].font = otherFont
		ws["K"+str(skuWriteRow)].alignment = centerAlign

		ws["N"+str(skuWriteRow)].value = str('*' + item[0] + '*') # Sku
		ws["N"+str(skuWriteRow)].font = otherFont
		ws["N"+str(skuWriteRow)].alignment = centerAlign

		# 13 line items print per page, after 13 lines are written we need to add an additional sidebar barcode for fulfillment bin
		# merge B:C, D:F, H:J, K:M, N:O
		# insert data B, D, G, H, K, N

		if ((skuWriteRow-5) % 14) == 0:
			sidebarMerge = ("A"+str(skuWriteRow)+":A"+str(skuWriteRow+13)+"")
			# print("A"+str(skuWriteRow)+":A"+str(skuWriteRow+13)+"")
			ws.merge_cells(range_string=sidebarMerge)
			ws["A"+str(skuWriteRow)].alignment = sidebarAlign
			ws["A"+str(skuWriteRow)].font = sideBarFont
			ws["A"+str(skuWriteRow)].value = shipBin

		skuWriteRow += 1
	wb.save(destinationName)
	wb.close()

def assignOrders(ordersListforAssign):
	print("INFO - Now assigning orders to warehouse 7 and 100")
	orderIds = []
	userId = ''
	countNJ = 0
	countNV = 0
	# values = {"orderIds":[],"userId":''}

	valuesNJ = {"orderIds":[],"userId":'0af695dc-65e8-414d-add3-0e314ecc152a'} # Warehouse 7 User (Currently CSS)
	# valuesNJ = {"orderIds":[],"userId":'4a21dd33-8ed1-4996-8eba-5c067a74ff76'} # Use For Testing - Assigns to Bill
	valuesNV = {"orderIds":[],"userId":'f7e760a2-048c-49f0-8d49-10fea8a5dce3'} # Warehouse 100 User (Currently Brent)
	# valuesNV = {"orderIds":[],"userId":'e68b8e34-54bc-41ec-b2aa-87f70145c600'} # Use For Testing - Assigns to James

	for order in ordersListforAssign:
		print('	INFO -	', order)
		# order[0] #ORDERID
		# order[1] #Warehouse Num
		if str(order[1]) == '7':
			valuesNJ["orderIds"].append(int(order[0]))
			countNJ += 1
			if countNJ % 100 == 0:
				print('INFO - Assigning', countNJ, 'orders to NJ')
				ShipStation.postShipStationHttpRequest('orders','assignuser',valuesNJ)
				print(valuesNJ)
				valuesNJ.clear()
				valuesNJ = {"orderIds":[],"userId":'0af695dc-65e8-414d-add3-0e314ecc152a'} # Warehouse 7 User (Currently CSS)
				print('INFO - Sleeping 1 seconds . . .')
				time.sleep(1)
				print('INFO - Done Sleeping')

		elif str(order[1]) == '100':
			valuesNV["orderIds"].append(int(order[0]))
			countNV += 1
			if countNV % 100 == 0:
				print('INFO - Assigning', countNV, 'orders to NV')
				ShipStation.postShipStationHttpRequest('orders','assignuser',valuesNV)
				print(valuesNV)
				valuesNV.clear()
				valuesNV = {"orderIds":[],"userId":'f7e760a2-048c-49f0-8d49-10fea8a5dce3'} # Warehouse 100 User (Currently Brent)
				print('INFO - Sleeping 1 seconds . . .')
				time.sleep(1)
				print('INFO - Done Sleeping')
	
	if valuesNJ["orderIds"]:
		print('INFO - Assigned', countNJ, 'orders to NJ')
		print(valuesNJ)
		ShipStation.postShipStationHttpRequest('orders','assignuser',valuesNJ)
		
	if valuesNV["orderIds"]:
		print('INFO - Assigned', countNV, 'orders to NV')
		print(valuesNV)
		ShipStation.postShipStationHttpRequest('orders','assignuser',valuesNV)

	valuesNJ.clear()
	valuesNV.clear()	

	valuesNJ = {"orderIds":[],"userId":'0af695dc-65e8-414d-add3-0e314ecc152a'} # Warehouse 7 User (Currently CSS)
	# valuesNJ = {"orderIds":[],"userId":'4a21dd33-8ed1-4996-8eba-5c067a74ff76'} # Use For Testing - Assigns to Bill
	valuesNV = {"orderIds":[],"userId":'f7e760a2-048c-49f0-8d49-10fea8a5dce3'} # Warehouse 100 User (Currently Brent)
	# valuesNV = {"orderIds":[],"userId":'e68b8e34-54bc-41ec-b2aa-87f70145c600'} # Use For Testing - Assigns to James

	countNJ = 0
	countNV = 0
	print("INFO - Orders have been assigned to warehouse 7 and 100\n")

	return;

main()