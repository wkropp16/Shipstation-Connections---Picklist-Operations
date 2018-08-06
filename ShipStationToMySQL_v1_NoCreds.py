import pypyodbc, ShipStation, datetime, json, time, MySQLdb, _mysql, pypyodbc, shutil
from datetime import datetime, timedelta
from openpyxl import load_workbook	#Import Openpyxl Library
from datetime import datetime as dt

def main():
	conn = MySQLdb.connect(host="localhost", user="REMOVED", passwd="REMOVED", db="shipstation", charset='utf8')
	cursor = conn.cursor()
	ordDate = (datetime.now()-timedelta(days=4)).strftime("%Y-%m-%d")
	shipDate = (datetime.now()-timedelta(days=6)).strftime("%Y-%m-%d")
	# ordDate = (datetime.now()-timedelta(days=6)).strftime("%Y-%m-%d")
	# shipDate = (datetime.now()-timedelta(days=20)).strftime("%Y-%m-%d")
	removeMergeIds = insertSS_Order(conn, cursor, ordDate)
	insertSS_Shipment(conn, cursor, shipDate)
	clearShippingBin(conn, cursor, removeMergeIds)
	print('')
	print("INFO - Closing Database Connection . . .")
	conn.commit()
	cursor.close()
	conn.close()
	print("INFO - Database Connection Closed . . .")
	print('')
	return;

def mailInnovationsPrices():
	rateTable = {}
	xlRT = ('../References/essentialvalues.xlsx') 							# Get Essential Values Workbook
	xlwbRT = load_workbook(filename = xlRT, data_only=True) 				# Load Essential Values
	miRateSheet = xlwbRT['MI_SHIPPING_TABLE'] 								# Find 'Rate' Sheet
	for row in miRateSheet.iter_rows(min_row=2, max_col=2, max_row=17):		#Make a loop for rows
		if row != "":
			col_a, col_b = row 						# splits the 'row' into two values col_A and col_B
			a_value = col_a.value 					# sets value of 'a' to be the value in the cell
			b_value = round(col_b.value, 2)			# sets value of 'b' to be a rounded (2decimal) value in the cell
			rateTable[a_value] = b_value 			# Adds to dictionary to look like {'Weight1': 'MonetaryFee1','Weight2': 'MonetaryFee2'}
	xlwbRT.close()
	return rateTable;

def getStoreIds():
	storeIds = {}
	xl = ('../References/essentialvalues.xlsx') 							# Get Workbook
	xlwb = load_workbook(filename = xl, data_only=True) 					# Load Workbook
	storeSetup = xlwb['STORE_SETUP'] 										# Find 'Rate' Sheet
	for row in storeSetup.iter_rows(min_row=2, max_col=2, max_row=60):		# Make a loop for rows
		if row != '':
			col_a, col_b = row
			a_value = str(col_a.value)		# Store ID
			b_value = col_b.value			# Store Name
			storeIds[a_value] = b_value
	xlwb.close()
	return storeIds;

def getStoreBins(): # Gets the stores from ../References/EssentialValues.xlsx to determine which stores should be on the picklist
	storeBins = {}
	xlfile = ('../References/essentialvalues.xlsx') 							# Specify File Name
	wb = load_workbook(filename = xlfile) 										# Define the 'workbook'
	EVPLI = wb['PICKLIST_INFO'] 												# Define Sheet
	for row in EVPLI.iter_rows(min_row=2, max_col=2, max_row=EVPLI.max_row):	# Make a loop for rows
		col_a, col_b = row
		if col_a.value != None:
			storeBins[col_a.value] = col_b.value
	wb.close()
	return storeBins;

def insertSS_Order(conn, cursor, ordDate):
	print('INFO - - - - - - - - - - - - - - - - - Processing Orders! ! !')
	# page = 11
	page = 1
	totalordercount = 0
	updateordercount = 0
	newordercount = 0
	ProductList = {}
	removeMergeIds = []
	removalChars = str.maketrans(dict.fromkeys('\',\",\\,`,/,&,%,’,,'))
	storeIds = getStoreIds()
	while True:
		params = ('orderDateStart=' + str(ordDate) + '&pageSize=500&page=' + str(page)) #Original Line
		# params = ('orderDateStart=2018-01-22&orderDateEnd=2018-03-18&pageSize=500&page=' + str(page)) #Modify this line for given circumstances
		ParseMe = json.loads(ShipStation.getShipStationHttpRequest('orders',params).content.decode('utf-8'))
		if ParseMe: print('INFO - JSON File Returned - Now Parsing Orders')
		thispage = ParseMe["page"]
		allpages = ParseMe["pages"]
		
		if int(thispage) > int(allpages):
			print('INFO - Breaking Loop')
			break

		elif int(thispage) <= int(allpages):
			start = time.time()
			print('INFO - Page ' + str(thispage) + ' of ' + str(allpages))
			print('INFO - Attempting append')
			for order2 in ParseMe["orders"]:

				###################################	order
				order = {}
				for key, value in order2.items(): # Check and modify JSON data to fit our need
					if ("None" or None or "none") in str(value): # Handle NULL data
						value = None
					if key.endswith('Date'): # Cut the Date string
						if not value == None:
							value = value[:10]
					if len(str(value)) > 45 and key in ('customerEmail', 'customerUsername'):
						value = value[:44]
					order[key] = value
				###################################	billTo
				billTo = {}
				if not order2['billTo'] == None:
					for key, value in order2['billTo'].items():
						if ("None" or None or "none") in str(value): # Handle NULL data
							value = None
						if ('\'' or '\"' or '&' or '\\' or '/' or ',' or '%' or '’' or '`') in str(value):
							value = value.translate(removalChars)
						if len(str(value)) > 44 and key in ('street1', 'street2', 'street3', 'name', 'company'):
							value = value[:44]
						billTo[key] = value
				###################################	shipTo
				shipTo = {}
				if not order2['shipTo'] == None:
					for key, value in order2['shipTo'].items():
						if ("None" or None or "none") in str(value): # Handle NULL data
							value = None
						if ('\'' or '\"' or '&' or '\\' or '/' or ',' or '%' or '’' or '`') in str(value):
							value = value.translate(removalChars)
						if len(str(value)) > 44 and key in ('street1', 'street2', 'street3', 'name', 'company'):
							value = value[:44]
						shipTo[key] = value
				###################################	items
				items = []
				if not order2['items'] == None:
					for product in order2['items']:
						p_product = {}
						for key,value in product.items():
							if ("None" or None or "none") in str(value): # Handle NULL data
								value = None
							if ('\'' or '\"' or '&' or '\\' or '/' or ',' or '%' or '’' or '`') in str(value) and not key in ("weight", "options"):
								value = value.translate(removalChars)
							if key.endswith('Date'): # Cut the Date string
								if not value == None:
									value = value[:10]
							if key == 'sku':
								if value is not None: 
									if (len(value) >= 45):
										value = value[:44]
							if key == 'name':
								if value is not None: 
									if (len(value) >= 45):
										value = value[:44]
							if not product['weight'] == None:
								for key1,value1 in product['weight'].items():
									if ("None" or None or "none") in str(value): # Handle NULL data
										value = None
									if key1 == "value":
										key1 = "weight_value"
									if key1 == "units":
										key1 = "weight_units"
									if key1 == "WeightUnits":
										key1 = "weight_WeightUnits"
									p_product[key1] = value1
							elif product['weight'] == None:
									p_product['weight_value'] = None
									p_product['weight_units'] = None
									p_product['weight_WeightUnits'] = None
							if not key in ("options", "weight"):
								p_product[key] = value
						items.append(p_product)
				###################################	weight
				weight = {}
				if not order2['weight'] == None:
					for key, value in order2['weight'].items():
						if ("None" or None or "none") in str(value): # Handle NULL data
							value = None
						if ('\'' or '\"' or '&' or '\\' or '/' or ',' or '%' or '’' or '`') in str(value):
							value = value.translate(removalChars)
						weight[key] = value
				###################################	dimensions
				dimensions = {}
				if not order2['dimensions'] == None:
					for key, value in order2['dimensions'].items():
						if ("None" or None or "none") in str(value): # Handle NULL data
							value = None
						if ('\'' or '\"' or '&' or '\\' or '/' or ',' or '%' or '’' or '`') in str(value):
							value = value.translate(removalChars)
						dimensions[key] = value
				###################################	insuranceOptions
				insuranceOptions = {}
				if not order2['insuranceOptions'] == None:
					for key, value in order2['insuranceOptions'].items():
						if ("None" or None or "none") in str(value): # Handle NULL data
							value = None
						if ('\'' or '\"' or '&' or '\\' or '/' or ',' or '%' or '’' or '`') in str(value):
							value = value.translate(removalChars)
						insuranceOptions[key] = value
				###################################	internationalOptions
				internationalOptions = {}
				customsItems = {}
				if not order2['internationalOptions'] == None:
					for key, value in order2['internationalOptions'].items():
						if ("None" or None or "none") in str(value): # Handle NULL data
							value = None
						if not order2['internationalOptions']['customsItems'] == None:
							for item in order2['internationalOptions']['customsItems']:
								for key1, value1 in item.items():
									if ("None" or None or "none") in str(value): # Handle NULL data
										value1 = None
									if ('\'' or '\"' or '&' or '\\' or '/' or ',' or '%' or '’' or '`') in str(value):
										value1 = str(value1).translate(removalChars)
						internationalOptions[key] = value
				###################################	advancedOptions
				advancedOptions = {}
				if not order2['advancedOptions'] == None:
					for key, value in order2['advancedOptions'].items():
						if ("None" or None or "none") in str(value): # Handle NULL data
							value = None
						if ('\'' or '\"' or '&' or '\\' or '/' or ',' or '%' or '’' or '`') in str(value):
							value = value.translate(removalChars)

						if key == "mergedIds":
							# if ("None" or None or "none" or "[]" or "[ ]") in str(value):
							if str(value) == '[]':
								value = ""
							else:	# Handle Merged Orders to add to database, SHOULD remove un-merged ID's further down the line.
								mergedOrderIds = ""
								first = True
								for mergedOrderId in value:
									print('INFO - MOIDs : ',mergedOrderId)
									removeMergeIds.append(mergedOrderId)
									print('INFO - RMOIDs : ',removeMergeIds)
									if first == True:
										mergedOrderIds = str(mergedOrderId)
										first = False
									else:
										mergedOrderIds = str(mergedOrderIds + ',' + str(mergedOrderId))
								value = mergedOrderIds


						if key == "storeId":
							if str(value) in storeIds:
								value = storeIds[str(value)]
								if str(order['tagIds']) == '[62652]':
									value = 'SP - Amazon FBA'
							else:
								value = str('NON-MAPPED ID - ' + str(value))

						advancedOptions[key] = value


				###################################	SQL Commands
				# CMD_GetColHeaders = 		('SELECT * FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME=ss_orders')
				CMD_CheckOrder = 			('SELECT orderId FROM ss_orders WHERE orderId=' + '\'' + str(order['orderId']) + '\'' + ';')
				CMD_CheckBillTo = 			('SELECT billTo_Name FROM ss_orders WHERE orderId=' + '\'' + str(order['orderId']) + '\'' + ';')
				CMD_CheckShipTo = 			('SELECT shipTo_Name FROM ss_orders WHERE orderId=' + '\'' + str(order['orderId']) + '\'' + ';')
				CMD_CheckAdvancedOptions = 	('SELECT ao_StoreId FROM ss_orders WHERE orderId=' + '\'' + str(order['orderId']) + '\'' + ';')

				CMD_InsertOrder = 			("INSERT INTO ss_orders (orderId, orderNumber, orderKey, orderStatus, orderDate, createDate, modifyDate, paymentDate, holdUntilDate, shipByDate, shipDate, userId, customerId, customerUsername, customerEmail, orderTotal, amountPaid, taxAmount, gift, requestedShippingService, carrierCode, serviceCode, packageCode, confirmation, externallyFulfilled, externallyFulfilledBy) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)")
				CMD_InsertBillTo = 			("UPDATE ss_orders SET billTo_Name=%s, billTo_Company=%s, billTo_Street1=%s, billTo_Street2=%s, billTo_Street3=%s, billTo_City=%s, billTo_State=%s, billTo_PostalCode=%s, billTo_Country=%s, billTo_Residential=%s WHERE orderId=%s")
				CMD_InsertShipTo = 			("UPDATE ss_orders SET shipTo_Name=%s, shipTo_Company=%s, shipTo_Street1=%s, shipTo_Street2=%s, shipTo_Street3=%s, shipTo_City=%s, shipTo_State=%s, shipTo_PostalCode=%s, shipTo_Country=%s, shipTo_Residential=%s WHERE orderId=%s")
				CMD_InsertAdvancedOptions = ("UPDATE ss_orders SET ao_SaturdayDelivery=%s, ao_MergedOrSplit=%s, ao_MergedIds=%s, ao_ParentId=%s, ao_StoreId=%s, ao_Source=%s, ao_billToParty=%s, ao_billToAccount=%s, ao_billToPostalCode=%s, ao_billToCountryCode=%s, ao_billToMyOtherAccount=%s WHERE orderId=%s")
				
				CMD_UpdateOrder = 			("UPDATE ss_orders SET orderNumber=%s, orderKey=%s, orderStatus=%s, orderDate=%s, createDate=%s, modifyDate=%s, paymentDate=%s, holdUntilDate=%s, shipByDate=%s, shipDate=%s, userId=%s, customerId=%s, customerUsername=%s, customerEmail=%s, orderTotal=%s, amountPaid=%s, taxAmount=%s, gift=%s, requestedShippingService=%s, carrierCode=%s, serviceCode=%s, packageCode=%s, confirmation=%s, externallyFulfilled=%s, externallyFulfilledBy=%s WHERE orderId=%s")
				CMD_UpdateAdvancedOptions = ("UPDATE ss_orders SET ao_SaturdayDelivery=%s, ao_MergedOrSplit=%s, ao_MergedIds=%s, ao_ParentId=%s, ao_StoreId=%s, ao_Source=%s, ao_billToParty=%s, ao_billToAccount=%s, ao_billToPostalCode=%s, ao_billToCountryCode=%s, ao_billToMyOtherAccount=%s WHERE orderId=%s")

				CMD_InsertSoldProduct = 	('INSERT INTO ss_order_items (orderId, orderItemId, lineItemKey, orderSku, orderSkuName, orderSkuQuantity, orderSkuUnitPrice, orderSkuTaxAmount, orderSkuShippingAmount, orderSkuProductId, orderSkuWeight, orderSkuWeightUnits, orderSkuCreateDate, orderSkuModifyDate) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)')
				CMD_CheckOrderItem = 		('SELECT lineItemKey FROM ss_order_items WHERE lineItemKey=%s AND orderId=%s')

				cursor.execute(CMD_CheckOrder)
				OrderExist = cursor.fetchone()
					
				if OrderExist == None: # Check if the order is in the database already
					print('INFO - ', advancedOptions['saturdayDelivery'], advancedOptions['mergedOrSplit'], advancedOptions['mergedIds'], advancedOptions['parentId'], advancedOptions['storeId'], advancedOptions['source'], advancedOptions['billToParty'], advancedOptions['billToAccount'], advancedOptions['billToPostalCode'], advancedOptions['billToCountryCode'], advancedOptions['billToMyOtherAccount'], order['orderId'])
					cursor.execute(CMD_InsertOrder, (order['orderId'], order['orderNumber'], order['orderKey'], order['orderStatus'], order['orderDate'], order['createDate'], order['modifyDate'], order['paymentDate'], order['holdUntilDate'], order['shipByDate'], order['shipDate'], order['userId'], order['customerId'], order['customerUsername'], order['customerEmail'], order['orderTotal'], order['amountPaid'], order['taxAmount'], order['gift'], order['requestedShippingService'], order['carrierCode'], order['serviceCode'], order['packageCode'], order['confirmation'], order['externallyFulfilled'], order['externallyFulfilledBy']))
					cursor.execute(CMD_InsertAdvancedOptions, (advancedOptions['saturdayDelivery'], advancedOptions['mergedOrSplit'], advancedOptions['mergedIds'], advancedOptions['parentId'], advancedOptions['storeId'], advancedOptions['source'], advancedOptions['billToParty'], advancedOptions['billToAccount'], advancedOptions['billToPostalCode'], advancedOptions['billToCountryCode'], advancedOptions['billToMyOtherAccount'], order['orderId']))
					cursor.execute(CMD_InsertShipTo, (shipTo['name'], shipTo['company'], shipTo['street1'], shipTo['street2'], shipTo['street3'], shipTo['city'], shipTo['state'], shipTo['postalCode'], shipTo['country'], shipTo['residential'], order['orderId']))
					cursor.execute(CMD_InsertBillTo, (billTo['name'], billTo['company'], billTo['street1'], billTo['street2'], billTo['street3'], billTo['city'], billTo['state'], billTo['postalCode'], billTo['country'], billTo['residential'], order['orderId']))
				else:
					cursor.execute(CMD_UpdateOrder, (order['orderNumber'], order['orderKey'], order['orderStatus'], order['orderDate'], order['createDate'], order['modifyDate'], order['paymentDate'], order['holdUntilDate'], order['shipByDate'], order['shipDate'], order['userId'], order['customerId'], order['customerUsername'], order['customerEmail'], order['orderTotal'], order['amountPaid'], order['taxAmount'], order['gift'], order['requestedShippingService'], order['carrierCode'], order['serviceCode'], order['packageCode'], order['confirmation'], order['externallyFulfilled'], order['externallyFulfilledBy'], order['orderId']))
					cursor.execute(CMD_InsertAdvancedOptions, (advancedOptions['saturdayDelivery'], advancedOptions['mergedOrSplit'], advancedOptions['mergedIds'], advancedOptions['parentId'], advancedOptions['storeId'], advancedOptions['source'], advancedOptions['billToParty'], advancedOptions['billToAccount'], advancedOptions['billToPostalCode'], advancedOptions['billToCountryCode'], advancedOptions['billToMyOtherAccount'], order['orderId']))
					print('INFO - Updated order: ' + str(order['orderId']))
				for product in items:
					cursor.execute(CMD_CheckOrderItem, (product['lineItemKey'], order['orderId']))
					ItemExist = cursor.fetchone()
					if ItemExist == None:
						cursor.execute(CMD_InsertSoldProduct, (order['orderId'], product['orderItemId'], product['lineItemKey'], product['sku'], product['name'], product['quantity'], product['unitPrice'], product['taxAmount'], product['shippingAmount'], product['productId'], product['weight_value'], product['weight_units'], product['createDate'], product['modifyDate']))

				totalordercount += 1
				if (totalordercount % 50) == 0:
					print('INFO - Orders checked: ', totalordercount)

		print('INFO - Total Orders Count: ' + str(totalordercount))
		end = time.time()
		print('INFO - Completed in: ' + (str(end - start))[:14] + ' seconds.')
		print('')
		conn.commit()
		if page % 25 == 0:
			print('INFO - Sleeping 1 minute . . .')
			time.sleep(60)
			print('INFO - Done Sleeping')
			print('')
		page += 1 #Add to pagination for loop
	conn.commit()
	print("INFO - Rows Commited to Database")
	print('')
	return removeMergeIds;

def insertSS_Shipment(conn, cursor, ordDate):
	print('INFO - - - - - - - - - - - - - - - - - Processing Shipments! ! !')
	rateTable = mailInnovationsPrices()
	oldTable = {1:2.64,2:2.64,3:2.64,4:2.64,5:2.77,6:2.77,7:2.77,8:2.77,9:3.34,10:3.42,11:3.56,12:3.69,13:3.84,14:3.96,15:4.10,16:4.22}
	page = 1
	while True:
		removalChars = str.maketrans(dict.fromkeys('\'\"\\/&,%'))
		params = ('createDateStart=' + str(ordDate) + '&pageSize=500&page=' + str(page)) # Original Line
		# params = ('createDateStart=2018-01-22&createDateEnd=2018-03-18&pageSize=500&page=' + str(page)) # Modify this line for given circumstances
		ParseMe = json.loads(ShipStation.getShipStationHttpRequest('shipments', params).content.decode('utf-8'))
		if ParseMe: print('INFO - JSON File Returned - Now Parsing Shipments')
		thispage = ParseMe["page"]
		allpages = ParseMe["pages"]
		
		if int(thispage) > int(allpages):
			print('INFO - Breaking Loop')
			break

		elif int(thispage) <= int(allpages):
			start = time.time()
			print('INFO - Page ' + str(thispage) + ' of ' + str(allpages))
			print('INFO - Attempting append')
			for shipment2 in ParseMe["shipments"]:
				shipment = {}
				for key, value in shipment2.items(): # Check and modify JSON data to fit our need
					if ("None" or None or "none") in str(value): # Handle NULL data
						value = None
					if key.endswith('Date'): # Cut the Date string
						if not value == None and len(value) > 11:
							value = value[:10]
					if key == "serviceCode" and value == "expedited_mail_innovations":
						shipmentCost = rateTable[shipment2['weight']['value']] # Current Table
						# shipmentCost = oldTable[shipment2['weight']['value']] # Old Rates (Ended 1/22/18)
						shipment['shipmentCost'] = shipmentCost

					if key == 'notifyErrorMessage' and not value == None:
						if len(value) >= 45:
							value = value[:44]

					if not key == "weight" and not key == "dimensions":
						shipment[key] = value

					if not shipment2['weight'] == None:
						for key1,value1 in shipment2['weight'].items():
							if ("None" or None or "none") in str(value): # Handle NULL data
								value = None
							if key1 == "value":
								key1 = "weight_value"
							if key1 == "units":
								key1 = "weight_units"
							if key1 == "WeightUnits":
								key1 = "weight_WeightUnits"
							shipment[key1] = value1
					elif shipment2['weight'] == None:
							shipment['weight_value'] = None
							shipment['weight_units'] = None
							shipment['weight_WeightUnits'] = None
					if not shipment2['dimensions'] == None:
						for key1,value1 in shipment2['dimensions'].items():
							if ("None" or None or "none") in str(value): # Handle NULL data
								value = None
							if key1 == "units":
								key1 = "dimensions_units"
							if key1 == "length":
								key1 = "dimensions_length"
							if key1 == "width":
								key1 = "dimensions_width"
							if key1 == "height":
								key1 = "dimensions_height"
							shipment[key1] = value1
					elif shipment2['dimensions'] == None:
							shipment['dimensions_units'] = None
							shipment['dimensions_length'] = None
							shipment['dimensions_width'] = None
							shipment['dimensions_height'] = None

				CMD_InsertShipment = ('INSERT INTO ss_shipments (shipmentId, orderId, orderNumber, orderKey, userId, createDate, shipDate, shipmentCost, insuranceCost, weight_value, weight_units, weight_WeightUnits, dimensions_units, dimensions_length, dimensions_width, dimensions_height, trackingNumber, isReturnLabel, carrierCode, serviceCode, packageCode, confirmation, warehouseId, voided, voidDate, marketplaceNotified, notifyErrorMessage) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)')
				CMD_UpdateShipment = ('UPDATE ss_shipments SET orderId=%s, orderNumber=%s, orderKey=%s, userId=%s, createDate=%s, shipDate=%s, shipmentCost=%s, insuranceCost=%s, weight_value=%s, weight_units=%s, weight_WeightUnits=%s, dimensions_units=%s, dimensions_length=%s, dimensions_width=%s, dimensions_height=%s, trackingNumber=%s, isReturnLabel=%s, carrierCode=%s, serviceCode=%s, packageCode=%s, confirmation=%s, warehouseId=%s, voided=%s, voidDate=%s, marketplaceNotified=%s, notifyErrorMessage=%s WHERE shipmentId=%s')
				CMD_CheckShipment = ('SELECT shipmentId FROM ss_shipments WHERE shipmentId=' + '\'' + str(shipment['shipmentId']) + '\'' + '')

				cursor.execute(CMD_CheckShipment)
				ShipmentExist = cursor.fetchone()
				if ShipmentExist == None: # Check if the order item is in the database already
					cursor.execute(CMD_InsertShipment, (shipment['shipmentId'], shipment['orderId'], shipment['orderNumber'], shipment['orderKey'], shipment['userId'], shipment['createDate'], shipment['shipDate'], shipment['shipmentCost'], shipment['insuranceCost'], shipment['weight_value'], shipment['weight_units'], shipment['weight_WeightUnits'], shipment['dimensions_units'], shipment['dimensions_length'], shipment['dimensions_width'], shipment['dimensions_height'], shipment['trackingNumber'], shipment['isReturnLabel'], shipment['carrierCode'], shipment['serviceCode'], shipment['packageCode'], shipment['confirmation'], shipment['warehouseId'], shipment['voided'], shipment['voidDate'], shipment['marketplaceNotified'], shipment['notifyErrorMessage']))
				else:
					cursor.execute(CMD_UpdateShipment, (shipment['orderId'], shipment['orderNumber'], shipment['orderKey'], shipment['userId'], shipment['createDate'], shipment['shipDate'], shipment['shipmentCost'], shipment['insuranceCost'], shipment['weight_value'], shipment['weight_units'], shipment['weight_WeightUnits'], shipment['dimensions_units'], shipment['dimensions_length'], shipment['dimensions_width'], shipment['dimensions_height'], shipment['trackingNumber'], shipment['isReturnLabel'], shipment['carrierCode'], shipment['serviceCode'], shipment['packageCode'], shipment['confirmation'], shipment['warehouseId'], shipment['voided'], shipment['voidDate'], shipment['marketplaceNotified'], shipment['notifyErrorMessage'], shipment['shipmentId']))
		end = time.time()
		print('INFO - Completed in: ' + str(end - start)[:14] + ' seconds.')
		print('')
		conn.commit()
		if page % 25 == 0:
			print('INFO - Sleeping 60 sec . . .')
			time.sleep(60)
			print('INFO - Done Sleeping')
			print('')
		page += 1 #Add to pagination for loop
		print('')
	conn.commit()
	print("INFO - Rows Commited to Database")
	print('')
	return;

def clearShippingBin(conn, cursor, removeMergeIds):
	print('\nINFO - Clearing the picklist from shipped orders! ! !')

	transcount = 0

	try:
		transname = ('../History_Scanforce_Transfers/IM4_' + str(dt.today().strftime('%Y%m%d_%H%M')) + '.CSS')
	except:
		transname = ('../History_Scanforce_Transfers/IM4_COUNDNTNAMEFILE.CSS')

	logfile = open(transname,'w')
	# logfile.write('RecType$|ItemCode$|WarehouseCode$|LotSerialNo$|BinLocation$|ToBinLocation$|UserCode$|TransactionDate$|QuantityTransfer\n')
	logfile.write('RecType$|TransactionDate$|ItemCode$|WarehouseCode$|ToWarehouseCode$|UnitOfMeasure$|BinLocation$|ToBinLocation$|LotSerialNo$|Valuation$|HeaderComment$|ForceCostType$|UserCode$|GLAccountKey$|JobCode$|CostCode$|CostType$|SalesOrderNo$|SaleOrderLineKey$|Quantity|SalesPrice|ItemCost|FreightCharge|Weight|UDF_SC_User_Code$|UDF_SC_Start_Time$|UDF_SC_Finish_Time$|UDF_SC_Duration$\n')

	binDate = dt.today().strftime('%Y%m%d')
	CMD_getOrdersFromPicklist = ('SELECT picklist.uniqueId, picklist.orderId, picklist.orderNumber, picklist.orderDate, picklist.skuIfMisc, picklist.orderSkuUnitPrice, ss_orders.ao_MergedIds FROM picklist INNER JOIN ss_orders on picklist.orderId=ss_orders.orderId')
	CMD_checkShipping = ('SELECT orderId, shipmentId FROM ss_shipments WHERE isReturnLabel=0 AND voided=0 AND orderId=%s')
	CMD_getInfoForScanForceFile = ('SELECT picklist.sku, picklist.quantity, picklist.warehouseAssigned, ss_orders.ao_StoreId, picklist.orderNumber, picklist.orderDate, picklist.skuIfMisc, picklist.orderSkuUnitPrice FROM picklist INNER JOIN ss_orders ON picklist.orderId=ss_orders.orderId WHERE picklist.orderId=%s')
	CMD_addToInvoiceTable = ('INSERT INTO invoice (orderId, orderNumber, orderDate, sku, skuIfMisc, orderSkuUnitPrice, quantity, status) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)')
	CMD_deleteFromPicklist = ('DELETE FROM picklist WHERE uniqueId=%s AND orderId=%s')
	CMD_getUID = ('SELECT uniqueId FROM picklist WHERE orderId=%s')

	storeBins = getStoreBins()

	cursor.execute(CMD_getOrdersFromPicklist)
	OrdersFromPicklist = cursor.fetchall()
	for order in OrdersFromPicklist: # [uniqueId, orderId, orderNumber, orderDate, skuIfMisc, orderSkuUnitPrice]
		print('INFO - Order In Picklist:', order)
		cursor.execute(CMD_checkShipping, (order[1],))
		checkShipping = cursor.fetchone()
		if checkShipping:
			cursor.execute(CMD_getInfoForScanForceFile, (order[1],))
			InfoForScanForceFile = cursor.fetchall()
			for ready2moveSKU in InfoForScanForceFile: # [picklist.sku, picklist.quantity, picklist.warehouseAssigned, ss_orders.ao_StoreId]
				print('INFO -	Ready to move bins:', ready2moveSKU)
				toWarehouse = '007'

				if ready2moveSKU[3] in storeBins:
					storeBin = storeBins[str(ready2moveSKU[3])]
				else:
					storeBin = ''

				if str(ready2moveSKU[2]) == '7':
					fromWarehouse = '007'
					logdata = '4|' + binDate + '|' + ready2moveSKU[0] + '|' + fromWarehouse + '|' + toWarehouse + '||E-FULFILLMENT|' + storeBin + '|||||BIL|||||||' + str(ready2moveSKU[1]) + '||||||||\n'
				elif str(ready2moveSKU[2]) == '16':
					fromWarehouse = '016'
					logdata = '4|' + binDate + '|' + ready2moveSKU[0] + '|' + fromWarehouse + '|' + toWarehouse + '||DIGITALDATA|' + storeBin + '|||||BIL|||||||' + str(ready2moveSKU[1]) + '||||||||\n'
				elif str(ready2moveSKU[2]) == '17':
					fromWarehouse = '007'
					logdata = '4|' + binDate + '|' + ready2moveSKU[0] + '|' + fromWarehouse + '|' + toWarehouse + '||E-FULFILLMENT|' + storeBin + '|||||BIL|||||||' + str(ready2moveSKU[1]) + '||||||||\n'
				elif str(ready2moveSKU[2]) == '100':
					fromWarehouse = '100'
					logdata = '4|' + binDate + '|' + ready2moveSKU[0] + '|' + fromWarehouse + '|' + toWarehouse + '||DOCK|' + storeBin + '|||||BIL|||||||' + str(ready2moveSKU[1]) + '||||||||\n'
				elif str(ready2moveSKU[2]) == '200':
					fromWarehouse = '200'
					logdata = '4|' + binDate + '|' + ready2moveSKU[0] + '|' + fromWarehouse + '|' + toWarehouse + '||DOCK|' + storeBin + '|||||BIL|||||||' + str(ready2moveSKU[1]) + '||||||||\n'
				else:
					fromWarehouse = str(ready2moveSKU[2])
					logdata = '4|' + binDate + '|' + ready2moveSKU[0] + '|' + fromWarehouse + '|' + toWarehouse + '||DOCK|' + storeBin + '|||||BIL|||||||' + str(ready2moveSKU[1]) + '||||||||\n'
				
				# logdata = ('4|' + ready2moveSKU[0] + '|' + fromWarehouse + '||SHIPPING|' + storeBin + '|BIL|' + binDate + '|' + str(ready2moveSKU[1]) + '\n')
				logfile.write(logdata)
				transcount += 1

				Ord_yyyymmdd = str(order[3]).translate(str.maketrans(dict.fromkeys('-')))

				cursor.execute(CMD_addToInvoiceTable, (order[1], order[2], Ord_yyyymmdd, ready2moveSKU[0], order[4], order[5], str(ready2moveSKU[1]), 'TransTo_StoreBin'))
				# 3596	409700104	73736	2018-05-04	G-9223		11.99	1	TransTo_StoreBin	INVOICENUMBER
				# UID	ORDERID		ORDNUM	ORDDATE		SKU			PPUNIT	QTY	INFO				INVNUM
				cursor.execute(CMD_deleteFromPicklist, (order[0], order[1]))
				# print("INFO - 		Deleted	| UID ", order[0], " | ORDNUM ", order[1], " | from Picklist")
				# print("INFO - 		Added	| UID ", order[0], " | ORDNUM ", order[1], " | to Invoice")

	for removalId in removeMergeIds: # [orderId, orderId, orderId]
		cursor.execute(CMD_getInfoForScanForceFile, (removalId,))
		InfoForScanForceFile = cursor.fetchall()
		for ready2moveSKU in InfoForScanForceFile: # [picklist.sku, picklist.quantity, picklist.warehouseAssigned, ss_orders.ao_StoreId]
			print('INFO - Ready to move bins:', ready2moveSKU, '	Merged ID')
			toWarehouse = '007'

			if ready2moveSKU[3] in storeBins:
				storeBin = storeBins[str(ready2moveSKU[3])]
			else:
				storeBin = ''

			if str(ready2moveSKU[2]) == '7':
				fromWarehouse = '007'
				logdata = '4|' + binDate + '|' + ready2moveSKU[0] + '|' + fromWarehouse + '|' + toWarehouse + '||E-FULFILLMENT|' + storeBin + '|||||BIL|||||||' + str(ready2moveSKU[1]) + '||||||||\n'
			elif str(ready2moveSKU[2]) == '16':
				fromWarehouse = '016'
				logdata = '4|' + binDate + '|' + ready2moveSKU[0] + '|' + fromWarehouse + '|' + toWarehouse + '||DIGITALDATA|' + storeBin + '|||||BIL|||||||' + str(ready2moveSKU[1]) + '||||||||\n'
			elif str(ready2moveSKU[2]) == '17':
				fromWarehouse = '007'
				logdata = '4|' + binDate + '|' + ready2moveSKU[0] + '|' + fromWarehouse + '|' + toWarehouse + '||E-FULFILLMENT|' + storeBin + '|||||BIL|||||||' + str(ready2moveSKU[1]) + '||||||||\n'
			elif str(ready2moveSKU[2]) == '100':
				fromWarehouse = '100'
				logdata = '4|' + binDate + '|' + ready2moveSKU[0] + '|' + fromWarehouse + '|' + toWarehouse + '||DOCK|' + storeBin + '|||||BIL|||||||' + str(ready2moveSKU[1]) + '||||||||\n'
			elif str(ready2moveSKU[2]) == '200':
				fromWarehouse = '200'
				logdata = '4|' + binDate + '|' + ready2moveSKU[0] + '|' + fromWarehouse + '|' + toWarehouse + '||DOCK|' + storeBin + '|||||BIL|||||||' + str(ready2moveSKU[1]) + '||||||||\n'
			else:
				fromWarehouse = str(ready2moveSKU[2])
				logdata = '4|' + binDate + '|' + ready2moveSKU[0] + '|' + fromWarehouse + '|' + toWarehouse + '||DOCK|' + storeBin + '|||||BIL|||||||' + str(ready2moveSKU[1]) + '||||||||\n'
			
			# logdata = ('4|' + ready2moveSKU[0] + '|' + fromWarehouse + '||SHIPPING|' + storeBin + '|BIL|' + binDate + '|' + str(ready2moveSKU[1]) + '\n')
			logfile.write(logdata)
			transcount += 1

			Ord_yyyymmdd = str(ready2moveSKU[5]).translate(str.maketrans(dict.fromkeys('-')))

			cursor.execute(CMD_addToInvoiceTable, (removalId, ready2moveSKU[4], Ord_yyyymmdd, ready2moveSKU[0], ready2moveSKU[6], ready2moveSKU[7], str(ready2moveSKU[1]), 'TransTo_StoreBin'))
			# 3596	409700104	73736	2018-05-04	G-9223		11.99	1	TransTo_StoreBin	INVOICENUMBER
			# UID	ORDERID		ORDNUM	ORDDATE		SKU			PPUNIT	QTY	INFO				INVNUM
			cursor.execute(CMD_getUID, (removalId,))
			getUID = cursor.fetchone()
			cursor.execute(CMD_deleteFromPicklist, (getUID, removalId))
			print("INFO - 		Deleted	| UID ", order[0], " | ORDNUM ", order[1], " | from Picklist")
			print("INFO - 		Added	| UID ", order[0], " | ORDNUM ", order[1], " | to Invoice")

	conn.commit()
	logfile.close()

	if transcount > 1:
		try:
			shutil.copy(transname, 'W:/ScanForce/Import/') # Un-Comment to auto move the transfer file to scanforce
			print('INFO - File moved to Import Folder for ScanForce')
		except:
			print('ERROR - Could not find drive 90w \"(\\\\cssdc1) W:\" ')
			shutil.copy(transname, '../History_Scanforce_Transfers/move_failures/')
			print('INFO - Moved to ../History_Scanforce_Transfers/move_failures/')
	print('INFO - Done Clearing Picklist! ! !\n')
	return;

main()

